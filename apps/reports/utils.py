from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from django.http import HttpResponse
from django.conf import settings
from datetime import datetime
import os

def get_company_settings():
    """Obtener configuración de la empresa"""
    try:
        from apps.core.models import CompanySettings
        return CompanySettings.get_settings()
    except:
        return None

def export_to_excel(data, headers, filename, sheet_name='Reporte', 
                    title='Reporte', include_totals=True, total_fields=None):
    """
    Exporta datos a un archivo Excel con formato profesional
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    money_format = '#,##0.00 "Bs."'
    number_format = '#,##0'
    
    # Obtener configuración de la empresa
    company = get_company_settings()
    company_name = company.company_name if company else 'M&S Melamina'
    nit = company.nit if company else '123456789'
    
    # ============ ENCABEZADO ============
    current_row = 1
    
    # Logo (usar logo rectangular)
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logos', 'logo.png')
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width = 200
            img.height = 80
            ws.add_image(img, 'A1')
            current_row = 4
        except:
            pass
    
    # Título de la empresa
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=company_name)
    cell.font = Font(bold=True, size=16, color='2C3E50')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # NIT
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=f'NIT: {nit}')
    cell.font = Font(size=10, color='7F8C8D')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # Título del reporte
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=title)
    cell.font = Font(bold=True, size=12, color='2C3E50')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # Fecha de generación
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    cell.font = Font(size=9, color='95A5A6')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2
    
    # ============ ENCABEZADOS DE TABLA ============
    for col, (field, title) in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    current_row += 1
    
    # ============ DATOS ============
    for row, item in enumerate(data, current_row):
        for col, (field, _) in enumerate(headers, 1):
            value = item.get(field, '')
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            
            # Formato según tipo de dato
            if isinstance(value, (int, float)):
                if 'precio' in field.lower() or 'total' in field.lower() or 'valor' in field.lower():
                    cell.number_format = money_format
                else:
                    cell.number_format = number_format
                cell.alignment = Alignment(horizontal='right')
            elif isinstance(value, str):
                cell.alignment = Alignment(horizontal='left')
    
    current_row = current_row + len(data)
    
    # ============ TOTALES ============
    if include_totals and total_fields and data:
        # Fila de totales
        for col, (field, title) in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value='')
            cell.border = border
            cell.font = Font(bold=True)
        
        # Calcular totales
        for field in total_fields:
            total = sum(item.get(field, 0) for item in data if isinstance(item.get(field, 0), (int, float)))
            for col, (f, title) in enumerate(headers, 1):
                if f == field:
                    cell = ws.cell(row=current_row, column=col, value=total)
                    cell.font = Font(bold=True, color='2C3E50')
                    cell.border = border
                    if 'precio' in field.lower() or 'total' in field.lower() or 'valor' in field.lower():
                        cell.number_format = money_format
                    else:
                        cell.number_format = number_format
                    cell.alignment = Alignment(horizontal='right')
                    break
        
        # Etiqueta de totales
        cell = ws.cell(row=current_row, column=1, value='TOTALES')
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # ============ AJUSTAR ANCHO DE COLUMNAS ============
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 25
    
    # ============ PIE DE PÁGINA ============
    footer_row = current_row + 2
    if company and company.receipt_footer:
        ws.merge_cells(f'A{footer_row}:{get_column_letter(len(headers))}{footer_row}')
        cell = ws.cell(row=footer_row, column=1, value=company.receipt_footer)
        cell.font = Font(size=8, color='95A5A6', italic=True)
        cell.alignment = Alignment(horizontal='center')
    
    # ============ CREAR RESPUESTA ============
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


def export_sales_report(sales):
    """Exportar reporte de ventas a Excel con formato profesional"""
    headers = [
        ('sale_number', 'N° Venta'),
        ('date', 'Fecha'),
        ('customer_name', 'Cliente'),
        ('user', 'Vendedor'),
        ('total', 'Total (Bs.)'),
        ('payment_method', 'Método de Pago'),
        ('status', 'Estado'),
    ]
    
    data = []
    for sale in sales:
        fecha_str = sale.date.strftime('%d/%m/%Y %H:%M')
        data.append({
            'sale_number': sale.sale_number,
            'date': fecha_str,
            'customer_name': sale.customer_name or 'Consumidor Final',
            'user': sale.user.username,
            'total': float(sale.total),
            'payment_method': sale.get_payment_method_display(),
            'status': sale.get_status_display(),
        })
    
    return export_to_excel(
        data=data,
        headers=headers,
        filename='reporte_ventas',
        sheet_name='Ventas',
        title='REPORTE DE VENTAS',
        include_totals=True,
        total_fields=['total']
    )


def export_products_report(products):
    """Exportar reporte de productos a Excel con formato profesional"""
    headers = [
        ('name', 'Producto'),
        ('barcode', 'Código de Barras'),
        ('category', 'Categoría'),
        ('group', 'Grupo'),
        ('sale_price', 'Precio Venta (Bs.)'),
        ('current_stock', 'Stock Actual'),
        ('unit', 'Unidad'),
        ('is_active', 'Estado'),
    ]
    
    data = []
    for product in products:
        data.append({
            'name': product.name,
            'barcode': product.barcode or '-',
            'category': product.category.name if product.category else 'Sin categoría',
            'group': product.group.name if product.group else 'Sin grupo',
            'sale_price': float(product.sale_price),
            'current_stock': float(product.current_stock),
            'unit': product.unit,
            'is_active': 'Activo' if product.is_active else 'Inactivo',
        })
    
    return export_to_excel(
        data=data,
        headers=headers,
        filename='reporte_productos',
        sheet_name='Productos',
        title='REPORTE DE PRODUCTOS',
        include_totals=True,
        total_fields=['sale_price', 'current_stock']
    )


def export_inventory_report(stocks):
    """Exportar reporte de inventario a Excel con formato profesional"""
    headers = [
        ('warehouse', 'Almacén'),
        ('product', 'Producto'),
        ('barcode', 'Código'),
        ('quantity', 'Cantidad'),
        ('unit', 'Unidad'),
        ('valor', 'Valor (Bs.)'),
    ]
    
    data = []
    for stock in stocks:
        valor = float(stock.quantity) * float(stock.product.purchase_price)
        data.append({
            'warehouse': stock.warehouse.name,
            'product': stock.product.name,
            'barcode': stock.product.barcode or '-',
            'quantity': float(stock.quantity),
            'unit': stock.product.unit,
            'valor': valor,
        })
    
    return export_to_excel(
        data=data,
        headers=headers,
        filename='reporte_inventario',
        sheet_name='Inventario',
        title='REPORTE DE INVENTARIO VALORADO',
        include_totals=True,
        total_fields=['quantity', 'valor']
    )